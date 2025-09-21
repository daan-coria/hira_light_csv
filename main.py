from pathlib import Path
import pandas as pd
import altair as alt
from datetime import datetime

from ingestion.excel_v010 import (
    load_staffing_rules_from_excel,
    load_resources_from_excel,
    load_shifts_from_excel,
)
from ingestion.census_loader import load_census_with_season
from logic.staffing_plan import build_staffing_plan_from_rules
from logic.position_control import build_position_control, compare_plan_vs_resources
from logic.scheduler import assign_staff_to_shifts
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# --- Styles ---
HEADER_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # light blue
HEADER_FONT = Font(bold=True, color="000000")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
OVERRIDE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # light green
NEGATIVE_FILL = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")  # light red
POSITIVE_FILL = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")  # green

FILENAME = "HIRA Light (IP) v0.12 (1).xlsx"


def _resolve_excel_path() -> Path:
    here = Path(__file__).resolve().parent
    root = here.parent
    candidates = [
        root / "data" / FILENAME,
        here / "data" / FILENAME,
        root / FILENAME,
    ]
    for p in candidates:
        if p.exists():
            return p
    tried = "\n - ".join(str(p) for p in candidates)
    raise FileNotFoundError(f"Could not find '{FILENAME}'. Tried:\n - {tried}")


def _format_sheet(ws, sheet_name=None):
    """Apply enhanced formatting with header colors, borders, and conditional fills."""

    # --- Header row ---
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN

    # --- Detect Gap column ---
    gap_idx = None
    if sheet_name == "Staffing vs Resources":
        for c_idx, header_cell in enumerate(ws[1], start=1):
            if str(header_cell.value).lower() in {"gap", "fte_gap", "shortage"}:
                gap_idx = c_idx
                break

    # --- Data rows ---
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = THIN_BORDER
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

            # Apply date formatting
            if "date" in str(ws.cell(row=1, column=cell.col_idx).value).lower():
                cell.number_format = "yyyy-mm-dd"

        # Highlight overrides
        if sheet_name == "Staffing Plan":
            for cell in row:
                if str(ws.cell(row=1, column=cell.col_idx).value).lower() == "census":
                    cell.fill = OVERRIDE_FILL

        # Highlight shortages/surpluses
        if sheet_name == "Staffing vs Resources" and gap_idx:
            gap_cell = row[gap_idx - 1]
            try:
                if gap_cell.value is not None:
                    gap_val = float(gap_cell.value)
                    if gap_val < 0:
                        gap_cell.fill = NEGATIVE_FILL
                    elif gap_val > 0:
                        gap_cell.fill = POSITIVE_FILL
            except Exception:
                pass

    # --- Column widths ---
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        header_val = str(ws.cell(row=1, column=col[0].column).value).lower()
        is_date_col = "date" in header_val

        max_length = max((len(str(c.value)) for c in col if c.value is not None), default=0)
        ws.column_dimensions[col_letter].width = max(15 if is_date_col else 10, max_length + 2)

    ws.freeze_panes = "A2"


def run_pipeline():
    excel_path = _resolve_excel_path()
    print(f"📘 Using workbook: {excel_path}")

    # --- Load inputs ---
    census_df = load_census_with_season(str(excel_path), sheet_name="Census Input")
    print("🌦️  Applied Nash seasonality from config/settings.yaml")

    rules_df = load_staffing_rules_from_excel(str(excel_path), sheet_name="Staffing Grid")
    resources_df = load_resources_from_excel(str(excel_path), sheet_name="Resource Input")

    shifts_df = load_shifts_from_excel(str(excel_path), sheet_name="Shifts Input")
    if shifts_df.empty:
        print("⚠️  No shifts found in Excel. Falling back to config/settings.yaml")
    else:
        print("⏰ Loaded shifts from Excel (Shifts Input sheet)")

    # --- Build staffing plan ---
    plan_df = build_staffing_plan_from_rules(census_df, rules_df)

    # --- Compare against resources ---
    pos_ctrl_df = build_position_control(resources_df)
    comp_df = compare_plan_vs_resources(plan_df, pos_ctrl_df)

    # --- Assign staff to shifts ---
    schedule_df = assign_staff_to_shifts(plan_df, shifts_df, yaml_path="config/settings.yaml")
    if not shifts_df.empty:
        print("✅ Assigned staff to Excel-defined shifts (ShiftSource=Excel in output)")
    else:
        print("✅ Assigned staff to YAML-defined shifts (ShiftSource=YAML in output)")

    # --- Normalize dates ---
    for df in [plan_df, comp_df, schedule_df]:
        if "Date" in df.columns:
            df["Date"] = pd.to_datetime(df["Date"]).dt.date

    # --- Season summary ---
    season_summary = (
        schedule_df
        .groupby(["SeasonLabel", "Role", "Shift"], as_index=False)
        .agg({"Assigned": "sum"})
        .sort_values(["SeasonLabel", "Role", "Shift"])
    )

    # --- Build Summary sheet ---
    summary_data = {
        "Run Date": [datetime.now().strftime("%Y-%m-%d")],
        "Input Workbook": [str(excel_path)],
        "Config File": ["config/settings.yaml"],
        "Staffing Plan Rows": [len(plan_df)],
        "Staffing vs Resources Rows": [len(comp_df)],
        "Staffing Schedule Rows": [len(schedule_df)],
        "Season Summary Rows": [len(season_summary)],
    }
    summary_df = pd.DataFrame(summary_data)

    # --- Write to Excel ---
    out_dir = excel_path.parent
    out_path = out_dir / "staffing_outputs.xlsx"

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        plan_df.to_excel(writer, sheet_name="Staffing Plan", index=False)
        comp_df.to_excel(writer, sheet_name="Staffing vs Resources", index=False)
        schedule_df.to_excel(writer, sheet_name="Staffing Schedule", index=False)
        season_summary.to_excel(writer, sheet_name="Summary by Season", index=False)

    # --- Apply formatting ---
    wb = load_workbook(out_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        _format_sheet(ws, sheet_name)
    wb.save(out_path)

    # --- Preview ---
    print("📝 Staffing Plan (preview):")
    print(plan_df.head(10))
    print("\n📊 Staffing vs Resources (preview):")
    print(comp_df.head(10))
    print("\n⏰ Staffing Schedule (preview):")
    print(schedule_df.head(10))


if __name__ == "__main__":
    run_pipeline()
